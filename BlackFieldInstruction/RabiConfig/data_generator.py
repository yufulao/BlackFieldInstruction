# ******************************************************************
#       /\ /|       @file       data_generator.py
#       \ V/        @brief      数据生成器
#       | "")       @author     Shadowrabbit, yingtu0401@gmail.com
#       /  |                    
#      /  \\        @Modified   2022/4/23
#    *(__\_\        @Copyright  Copyright (c) 2022, Shadowrabbit
# ******************************************************************
# !/usr/bin/env python
# -*- coding: utf-8 -*-
import codecs
import os
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import str_util
from config_def import MANIFEST_PATH, XLSX_FOLDER, DATA_TARGET_FOLDER
from str_util import fix_str


class DataGenerator:
    # brief
    # private
    def __init__(self):
        # 需要生成数据的excel文件名清单
        self._excel_set = set()
        if not os.path.exists(MANIFEST_PATH):
            self._manifest_file = codecs.open(MANIFEST_PATH, "a", "utf-8")
            self._manifest_file.write("")
            self._manifest_file.close()
        self._manifest_file = codecs.open(MANIFEST_PATH, "r", "utf-8")  # 清单文件 读权限

    # brief
    # private
    def __del__(self):
        self._manifest_file.close()

    # brief 生成数据
    # public
    def generate_data(self):
        print("开始生成数据")
        self.check_manifest()
        self.generate_data_by_folder()

    # brief 检查清单文件
    # private
    def check_manifest(self):
        # 获取清单文件中的excel文件名
        self._excel_set = set()
        for line in self._manifest_file:
            self._excel_set.add(line.strip())

    # brief 通过目录生成数据文件
    # private
    def generate_data_by_folder(self):
        for dir_path, dir_names, file_names in os.walk(XLSX_FOLDER):
            # 过滤掉清单中不存在的文件
            valid_filename_list = list(filter(self.file_filter, file_names))
            # 生成数据文件
            folder = dir_path.replace('\\', '/')
            for filename in valid_filename_list:
                self.generate_datafile_by_path(f"{folder}/{filename}")

    # brief 通过路径生成数据文件
    # private
    def generate_datafile_by_path(self, filepath):
        # 工作簿实例
        workbook = load_workbook(filepath, data_only=True)
        if not workbook:
            print(f"工作簿加载异常:{filepath}")
            return
        sheet_name_list = workbook.get_sheet_names()
        if not sheet_name_list:
            print(f"找不到任何表格:{filepath}")
            return
        # 清单表
        manifest_sheet: Worksheet = workbook.get_sheet_by_name(sheet_name_list[0])
        # 遍历清单工作簿
        for i in range(manifest_sheet.max_row):
            # 清单第一列记录需要生成数据的表格
            sheet_name = str(manifest_sheet.cell(column=1, row=i + 1).value)
            fix_sheet_name = str_util.fix_str(sheet_name)
            if not fix_sheet_name or fix_sheet_name == "":
                print(f"找不到sheet名:{filepath} {fix_sheet_name}")
                continue
            # 数据表
            data_sheet = workbook.get_sheet_by_name(sheet_name)
            if not data_sheet:
                print(f"sheet不存在:{filepath} {fix_sheet_name}")
                continue
            # 数据文件路径
            output_path = f"{DATA_TARGET_FOLDER}/Cfg{sheet_name}.txt"
            self.generate_datafile_by_sheet(data_sheet, output_path)

    # brief 通过表名生成数据文件
    # private
    @staticmethod
    def generate_datafile_by_sheet(data_sheet: Worksheet, output_path):
        print(output_path)
        data_file = codecs.open(output_path, "w", "utf-8")
        for row in range(data_sheet.max_row):
            # 当前行数据列表
            data_row_list = []
            for col in range(data_sheet.max_column):
                value = str(data_sheet.cell(column=col + 1, row=row + 1).value)
                # 前三行特殊规则 过滤掉特殊字符
                if row < 3:
                    value = fix_str(value)
                # 空值无视
                if not value:
                    continue
                data_row_list.append(value)
            data_file.write("#".join(data_row_list))
            data_file.write(os.linesep)
        data_file.close()

    # brief 文件名是否在清单中
    # private
    def file_filter(self, filename) -> bool:
        return filename in self._excel_set
