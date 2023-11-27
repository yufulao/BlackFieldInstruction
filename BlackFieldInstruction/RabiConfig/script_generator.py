# ******************************************************************
#       /\ /|       @file       script_generator.py
#       \ V/        @brief      解析脚本生成器
#       | "")       @author     Shadowrabbit, yingtu0401@gmail.com
#       /  |                    
#      /  \\        @Modified   2022/4/23
#    *(__\_\        @Copyright  Copyright (c) 2022, Shadowrabbit
# ******************************************************************
# !/usr/bin/env python
# -*- coding: utf-8 -*-
import codecs
import os

import config_def
import field_data
import field_info

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import str_util
from config_def import XLSX_FOLDER, MANIFEST_PATH, SCRIPTS_TARGET_FOLDER, DATA_TARGET_FOLDER
from str_util import to_big_camel_case


class ScriptGenerator:
    # brief
    # private
    def __init__(self):
        self._temp_field_type = ""  # 字段类型
        self._temp_annotate = ""  # 字段注释
        self._temp_field_name = ""  # 字段名称
        self._excel_set = set()
        self._temp_base_field_info_map = {}  # 当前表内 非自定义结构的字段 key:field_name
        self._temp_custom_field_info_map = {}  # 当前表内 自定义结构的字段
        self._temp_field_info_list = []  # 有序 所有字段信息列表
        # 需要生成数据的excel文件名清单
        if not os.path.exists(MANIFEST_PATH):
            self._manifest_file = codecs.open(MANIFEST_PATH, "a", "utf-8")
            self._manifest_file.write("")
            self._manifest_file.close()
        self._manifest_file = codecs.open(MANIFEST_PATH, "r", "utf-8")  # 清单文件 读权限

    # brief
    # private
    def __del__(self):
        self._manifest_file.close()

    # brief 生成解析代码
    # public
    def generate_script(self):
        print("开始生成脚本")
        self.check_manifest()
        self.generate_script_by_folder()

    # brief 检查清单文件
    # private
    def check_manifest(self):
        # 获取清单文件中的excel文件名
        self._excel_set = set()
        for line in self._manifest_file:
            self._excel_set.add(line.strip())

    # brief 根据目录生成cs代码
    # private
    def generate_script_by_folder(self):
        # 遍历excel目标下全部excel文件(包含子目录)
        for dir_path, dir_names, file_names in os.walk(XLSX_FOLDER):
            # 过滤掉清单中不存在的文件
            valid_filename_list = list(filter(self.file_filter, file_names))
            # 当前目录
            folder = dir_path.replace('\\', '/')
            # 生成数据文件
            for filename in valid_filename_list:
                self.generate_script_by_path(f"{folder}/{filename}")

    # brief 通过路径生成cs代码
    # private
    # param filepath excel文件的路径
    def generate_script_by_path(self, filepath):
        # 工作簿实例
        workbook = load_workbook(filepath, data_only=True)
        if not workbook:
            print(f"工作簿加载异常:{filepath}")
            return
        sheet_name_list = workbook.get_sheet_names()
        if not sheet_name_list:
            print(f"找不到任何表格:{filepath}")
            return
        # 清单表
        manifest_sheet: Worksheet = workbook.get_sheet_by_name(sheet_name_list[0])
        # 遍历清单工作簿
        for i in range(manifest_sheet.max_row):
            # 清单第一列记录需要生成数据的表格名字
            sheet_name = str(manifest_sheet.cell(column=1, row=i + 1).value)
            fix_sheet_name = str_util.fix_str(sheet_name)
            if not fix_sheet_name or fix_sheet_name == "":
                print(f"找不到sheet名:{filepath} {fix_sheet_name}")
                continue
            # 数据表
            data_sheet = workbook.get_sheet_by_name(sheet_name)
            if not data_sheet:
                print(f"sheet不存在:{filepath} {fix_sheet_name}")
                continue
            # 数据文件路径
            output_path = f"{SCRIPTS_TARGET_FOLDER}/Cfg{sheet_name}.cs"
            self.generate_script_by_sheet(data_sheet, output_path, sheet_name, filepath)

    # brief 通过表格生成cs代码
    # private
    def generate_script_by_sheet(self, data_sheet: Worksheet, output_path, sheet_name, filepath):
        print(output_path)
        class_name = f"Cfg{sheet_name}"
        script_file = codecs.open(output_path, "w", "utf-8")
        # 清空结构
        self._temp_base_field_info_map.clear()
        self._temp_custom_field_info_map.clear()
        self._temp_field_info_list.clear()
        # 结构分析
        self.parse_data_struct(data_sheet)
        # 写入标题
        self.write_header_str(script_file, class_name, filepath)
        # 写入库引用
        self.write_lib(script_file)
        # 写入代码
        self.write_code(script_file, class_name)
        script_file.close()

    # brief 解析表格字段数据结构
    # private
    def parse_data_struct(self, data_sheet: Worksheet):
        # 遍历每一列
        for col in range(data_sheet.max_column):
            # 每一列的前三行做特殊字段解析
            for row in range(3):
                value = str(data_sheet.cell(column=col + 1, row=row + 1).value)
                # 每列的第1行 字段名
                if row == 0:
                    self._temp_field_name = value
                    continue
                # 每列的第2行 注释
                elif row == 1:
                    self._temp_annotate = value
                    continue
                # 每列的第3行 字段数据类型
                elif row == 2:
                    self._temp_field_type = value
                    continue
            # 过滤空值
            if self._temp_field_name == "None":
                continue
            # 根据当前field data分析field info
            self.field_analyze()

    # brief field结构分析
    # private
    def field_analyze(self):
        # 当前field data是否是自定义类型数据结构的一部分
        is_custom_type = self.is_custom_type(self._temp_field_name)
        # 非自定义结构的情况
        if not is_custom_type:
            # 当前字段数据
            cur_field_data = field_data.FieldData(self._temp_field_name, self._temp_annotate, self._temp_field_type)
            # 当前字段没有被记录的情况 创建新新的field info
            if self._temp_field_name not in self._temp_base_field_info_map:
                cur_field_info = field_info.FieldInfo("")
                self._temp_base_field_info_map[self._temp_field_name] = cur_field_info
                # 尝试加入当前field data
                self._temp_field_info_list.append(cur_field_info)
            self._temp_base_field_info_map[self._temp_field_name].try_add_field(cur_field_data)
            return
        # 自定义结构体的情况
        struct_name, field_name = self.get_struct_and_field_name(self._temp_field_name)
        cur_field_data = field_data.FieldData(field_name, self._temp_annotate, self._temp_field_type)
        # 当前没被记录的结构体
        if struct_name not in self._temp_custom_field_info_map:
            cur_field_info = field_info.FieldInfo(struct_name)
            self._temp_custom_field_info_map[struct_name] = cur_field_info
            self._temp_field_info_list.append(cur_field_info)
        self._temp_custom_field_info_map[struct_name].try_add_field(cur_field_data)

    # brief 文件名是否在清单中
    # private
    def file_filter(self, filename) -> bool:
        return filename in self._excel_set

    # brief 当前数据类型是否是自定义结构体
    # private
    @staticmethod
    def is_custom_type(field_name: str):
        return field_name.find("-") > 0

    # brief 获取结构体名称 当前结构体内字段名称
    # private
    @staticmethod
    def get_struct_and_field_name(field_name: str):
        struct_name_str = field_name.split("-")
        if len(struct_name_str) < 2:
            print(f"无法正确识别结构体:{field_name}")
            return
        return struct_name_str[0], struct_name_str[1]

    # brief 写入标题字符
    # private
    @staticmethod
    def write_header_str(script_file, filename, filepath):
        if not isinstance(filename, str):
            print(f"文件名格式错误: {filename}")
            return
        if not script_file:
            return
        str_list = [
            "// ******************************************************************",
            f"//       /\\ /|       @file       {filename}.cs",
            f"//       \\ V/        @brief      excel数据解析(由python自动生成) {filepath}",
            "//       | \"\")       @author     Shadowrabbit, yingtu0401@gmail.com",
            "//       /  |",
            "//      /  \\\\        @Modified   2022-04-25 13:25:11",
            "//    *(__\\_\\        @Copyright  Copyright (c)  2022, Shadowrabbit",
            "// ******************************************************************"
        ]
        result = os.linesep.join(str_list)
        script_file.write(result)

    # brief 写入库引用
    # private
    @staticmethod
    def write_lib(script_file):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write("using UnityEngine;")
        script_file.write(os.linesep)
        script_file.write("using System;")
        script_file.write(os.linesep)
        script_file.write("using System.Linq;")
        script_file.write(os.linesep)
        script_file.write("using System.Collections.Generic;")

    # brief 写入代码
    # private
    def write_code(self, script_file, class_name):
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write(f"namespace {config_def.NAME_SPACE_NAME}")
        script_file.write(os.linesep)
        script_file.write("{")
        # 写入行数据的类定义
        self.write_row_data_class(script_file, class_name)
        # 写入自定义数据的类定义
        self.write_custom_structs_def(script_file)
        # 写入类
        self.write_class(script_file, class_name)
        script_file.write(os.linesep)
        script_file.write("}")

    # brief 写入每行数据的类定义
    # private
    def write_row_data_class(self, script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(f"    public class Row{class_name}")
        script_file.write(os.linesep)
        script_file.write("    {")
        for info in self._temp_field_info_list:
            script_file.write(os.linesep)
            script_file.write(
                f"        public {info.get_field_type()} {info.get_field_name()}; //{info.get_field_annotate()}")
        script_file.write(os.linesep)
        script_file.write("    }")

    # brief 写入自定义数据的类定义
    # private
    def write_custom_structs_def(self, script_file):
        if not script_file:
            return
        for custom_field in self._temp_custom_field_info_map.values():
            script_file.write(os.linesep)
            script_file.write(os.linesep)
            script_file.write(f"    public struct {custom_field.parent_class_name}")
            script_file.write(os.linesep)
            script_file.write("    {")
            for data in custom_field.field_data_list:
                script_file.write(os.linesep)
                script_file.write(
                    f"        public {data.field_type} {data.field_name}; //{data.annotate}")
            script_file.write(os.linesep)
            script_file.write("    }")

    # brief 写入类
    # private
    def write_class(self, script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write(f"    public class {class_name}")
        script_file.write(os.linesep)
        script_file.write("    {")
        self.write_fields(script_file, f"Row{class_name}")
        self.write_function_find(script_file, class_name)
        self.write_function_load(script_file, class_name)
        self.write_function_group_find(script_file, class_name)
        self.write_function_parse_row(script_file, class_name)
        script_file.write(os.linesep)
        script_file.write("    }")

    # brief 写入字段定义
    # private
    def write_fields(self, script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(
            f"        private readonly Dictionary<string, {class_name}> _configs "
            f"= new Dictionary<string, {class_name}>(); //cfgId映射row")
        # 带有组标记的字段 额外创建一个字典用于分组
        for info in self._temp_base_field_info_map.values():
            if info.is_group_field():
                script_file.write(os.linesep)
                script_file.write(
                    f"        private readonly Dictionary<{info.get_field_type()}, List<{class_name}>> "
                    f"_{info.get_field_name()}ConfigGroup = "
                    f"new Dictionary<{info.get_field_type()}, List<{class_name}>>();")
        script_file.write(os.linesep)
        script_file.write(
            f"        public {class_name} this[string key] => _configs.ContainsKey(key) "
            f"? _configs[key] : throw new Exception($\"找不到配置 Cfg:{{GetType()}} key:{{key}}\");")
        script_file.write(os.linesep)
        script_file.write(
            f"        public {class_name} this[int id] => _configs.ContainsKey(id.ToString()) "
            f"? _configs[id.ToString()] : throw new Exception($\"找不到配置 Cfg:{{GetType()}} key:{{id}}\");")
        script_file.write(os.linesep)
        script_file.write(f"        public List<{class_name}> AllConfigs => _configs.Values.ToList();")

    # brief 写入查找函数
    # private
    @staticmethod
    def write_function_find(script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write("        /// <summary>")
        script_file.write(os.linesep)
        script_file.write("        /// 获取行数据")
        script_file.write(os.linesep)
        script_file.write("        /// </summary>")
        script_file.write(os.linesep)
        script_file.write(f"        public Row{class_name} Find(int i)")
        script_file.write(os.linesep)
        script_file.write("        {")
        script_file.write(os.linesep)
        script_file.write("            return this[i];")
        script_file.write(os.linesep)
        script_file.write("        }")
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write("        /// <summary>")
        script_file.write(os.linesep)
        script_file.write("        /// 获取行数据")
        script_file.write(os.linesep)
        script_file.write("        /// </summary>")
        script_file.write(os.linesep)
        script_file.write(f"        public Row{class_name} Find(string i)")
        script_file.write(os.linesep)
        script_file.write("        {")
        script_file.write(os.linesep)
        script_file.write("            return this[i];")
        script_file.write(os.linesep)
        script_file.write("        }")

    # brief 写入加载函数
    # private
    def write_function_load(self, script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write("        /// <summary>")
        script_file.write(os.linesep)
        script_file.write("        /// 加载表数据")
        script_file.write(os.linesep)
        script_file.write("        /// </summary>")
        script_file.write(os.linesep)
        script_file.write("        public void Load()")
        script_file.write(os.linesep)
        script_file.write("        {")
        script_file.write(os.linesep)
        script_file.write("            var reader = new CsvReader();")
        script_file.write(os.linesep)
        # 切片截取Assets后的目录
        start_index = DATA_TARGET_FOLDER.find("Assets")
        data_path = DATA_TARGET_FOLDER[start_index:]
        script_file.write(f"            reader.LoadText(\"{data_path}/{class_name}.txt\", 3);")
        script_file.write(os.linesep)
        script_file.write("            var rows = reader.GetRowCount();")
        script_file.write(os.linesep)
        script_file.write("            for (var i = 0; i < rows; ++i)")
        script_file.write(os.linesep)
        script_file.write("            {")
        script_file.write(os.linesep)
        script_file.write("                var row = reader.GetColValueArray(i);")
        script_file.write(os.linesep)
        script_file.write("                var data = ParseRow(row);")
        script_file.write(os.linesep)
        script_file.write("                if (!_configs.ContainsKey(data.key))")
        script_file.write(os.linesep)
        script_file.write("                {")
        script_file.write(os.linesep)
        script_file.write("                    _configs.Add(data.key, data);")
        script_file.write(os.linesep)
        script_file.write("                }")
        for info in self._temp_base_field_info_map.values():
            if info.is_group_field():
                field_name = info.get_field_name()
                script_file.write(os.linesep)
                script_file.write(os.linesep)
                script_file.write(f"                if (!_{field_name}ConfigGroup.ContainsKey(data.{field_name}))")
                script_file.write(os.linesep)
                script_file.write("                {")
                script_file.write(os.linesep)
                script_file.write(f"                     _{field_name}ConfigGroup.Add"
                                  f"(data.{field_name}, new List<Row{class_name}>());")
                script_file.write(os.linesep)
                script_file.write("                }")
                script_file.write(os.linesep)
                script_file.write(os.linesep)
                script_file.write(f"                _{field_name}ConfigGroup[data.{field_name}].Add(data);")
        script_file.write(os.linesep)
        script_file.write("            }")
        script_file.write(os.linesep)
        script_file.write("        }")

    # brief 写入组查询函数
    # private
    def write_function_group_find(self, script_file, class_name):
        if not script_file:
            return
        for info in self._temp_base_field_info_map.values():
            if info.is_group_field():
                field_name = info.get_field_name()
                script_file.write(os.linesep)
                script_file.write(os.linesep)
                script_file.write("        /// <summary>")
                script_file.write(os.linesep)
                script_file.write(f"        /// 根据{field_name}值获取分组")
                script_file.write(os.linesep)
                script_file.write("        /// </summary>")
                script_file.write(os.linesep)
                script_file.write(f"        public List<Row{class_name}> "
                                  f"GetListBy{to_big_camel_case(field_name)}"
                                  f"({info.get_field_type()} groupValue)")
                script_file.write(os.linesep)
                script_file.write("        {")
                script_file.write(os.linesep)
                script_file.write(f"                return _{field_name}ConfigGroup.ContainsKey(groupValue) "
                                  f"? _{field_name}ConfigGroup[groupValue] "
                                  ": throw new Exception($\"找不到组 Cfg:{GetType()} groupId:{groupValue}\");")
                script_file.write(os.linesep)
                script_file.write("        }")

    # brief 写入行解析函数
    # private
    def write_function_parse_row(self, script_file, class_name):
        if not script_file:
            return
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write("        /// <summary>")
        script_file.write(os.linesep)
        script_file.write("        /// 解析行")
        script_file.write(os.linesep)
        script_file.write("        /// </summary>")
        script_file.write(os.linesep)
        script_file.write("        /// <param name=\"col\"></param>")
        script_file.write(os.linesep)
        script_file.write("        /// <returns></returns>")
        script_file.write(os.linesep)
        script_file.write(f"        private Row{class_name} ParseRow(string[] col)")
        script_file.write(os.linesep)
        script_file.write("        {")
        script_file.write(os.linesep)
        script_file.write("            //列越界")
        script_file.write(os.linesep)
        # 数据列长度
        max_col = 0
        for info in self._temp_field_info_list:
            max_col += info.appear_count
        script_file.write(f"            if (col.Length < {max_col})")
        script_file.write(os.linesep)
        script_file.write("            {")
        script_file.write(os.linesep)
        script_file.write("                Debug.LogError($\"配置表字段行数越界:{GetType()}\");")
        script_file.write(os.linesep)
        script_file.write("                return null;")
        script_file.write(os.linesep)
        script_file.write("            }")
        script_file.write(os.linesep)
        script_file.write(os.linesep)
        script_file.write(f"            var data = new Row{class_name}();")
        script_file.write(os.linesep)
        script_file.write("            var rowHelper = new RowHelper(col);")
        for info in self._temp_field_info_list:
            # 基础类型数据的情况
            if not info.parent_class_name and info.array_length <= 1:
                self.write_base_type(script_file, info)
                continue
            # 自定义类型的情况
            if info.parent_class_name and info.array_length <= 1:
                self.write_custom_type(script_file, info)
                continue
            # 基础类型数组的情况
            if not info.parent_class_name and info.array_length > 1:
                self.write_base_type_array(script_file, info)
                continue
            # 自定义类型数组的情况
            else:
                self.write_custom_type_array(script_file, info)
        script_file.write(os.linesep)
        script_file.write(f"            return data;")
        script_file.write(os.linesep)
        script_file.write("        }")

    # brief 写入基础类型数据
    # private
    @staticmethod
    def write_base_type(script_file, info):
        script_file.write(os.linesep)
        script_file.write(f"            data.{info.get_field_name()} = "
                          f"CsvUtility.To{info.get_field_type().capitalize()}(rowHelper.ReadNextCol()); "
                          f"//{info.get_field_annotate()}")

    # brief 写入自定义类型数据
    # private
    @staticmethod
    def write_custom_type(script_file, info):
        script_file.write(os.linesep)
        # 遍历自定义结构每个字段
        for data in info.field_data_list:
            script_file.write(f"            data.{info.get_field_name()}.{data.field_name} = "
                              f"CsvUtility.To{data.field_type.capitalize()}(rowHelper.ReadNextCol()); "
                              f"//{data.annotate}")

    # brief 写入基础类型数组
    # private
    @staticmethod
    def write_base_type_array(script_file, info):
        script_file.write(os.linesep)
        field_type = info.get_field_type()
        sub_type = field_type[:len(field_type) - 2]
        script_file.write(f"            data.{info.get_field_name()} = new {sub_type}[{info.array_length}];")
        for i in range(info.array_length):
            script_file.write(os.linesep)
            # string[]
            array_type = info.get_field_type()
            # string
            sub_type = array_type[:len(array_type) - 2]
            script_file.write(f"            data.{info.get_field_name()}[{i}] = "
                              f"CsvUtility.To{sub_type.capitalize()}(rowHelper.ReadNextCol()); "
                              f"//{info.get_field_annotate()}")

    # brief 写入自定义类型数组
    # private
    @staticmethod
    def write_custom_type_array(script_file, info):
        script_file.write(os.linesep)
        script_file.write(
            f"            data.{info.get_field_name()} = new {info.parent_class_name}[{info.array_length}];")
        for i in range(info.appear_count):
            array_index = i // len(info.field_data_list)  # 当前是数组中第几个结构体
            field_index = i % len(info.field_data_list)  # 结构体中的第几个参数
            # 数组元素更换 分配内存
            if field_index == 0:
                script_file.write(os.linesep)
                script_file.write(f"            data.{info.get_field_name()}[{array_index}] = "
                                  f"new {info.parent_class_name}();")
            script_file.write(os.linesep)
            data = info.field_data_list[field_index]
            script_file.write(f"            data.{info.get_field_name()}[{array_index}].{data.field_name} = "
                              f"CsvUtility.To{data.field_type.capitalize()}(rowHelper.ReadNextCol()); "
                              f"//{data.annotate}")
