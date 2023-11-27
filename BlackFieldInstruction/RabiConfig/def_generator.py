# ******************************************************************
#       /\ /|       @file       def_generator.py
#       \ V/        @brief      定义生成器
#       | "")       @author     Shadowrabbit, yingtu0401@gmail.com
#       /  |                    
#      /  \\        @Modified   2022/5/2
#    *(__\_\        @Copyright  Copyright (c) 2022, Shadowrabbit
# ******************************************************************
# !/usr/bin/env python
# -*- coding: utf-8 -*-
import codecs
import os
import config_def
import str_util

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from config_def import MANIFEST_PATH, DEF_TARGET_FOLDER, XLSX_FOLDER, DEF_MARK, ANNOTATE_MARK
from str_util import to_big_camel_case


class DefGenerator:
    # brief
    # private
    def __init__(self):
        # 需要生成数据的excel文件名清单
        self._excel_set = set()
        if not os.path.exists(MANIFEST_PATH):
            self._manifest_file = codecs.open(MANIFEST_PATH, "a", "utf-8")
            self._manifest_file.write("")
            self._manifest_file.close()
        self._manifest_file = codecs.open(MANIFEST_PATH, "r", "utf-8")  # 清单文件 读权限
        return

    # brief
    # private
    def __del__(self):
        self._manifest_file.close()
        return

    # brief 生成
    # public
    def generate_def(self):
        print("开始生成")
        self.check_manifest()
        self.generate_def_by_folder()

    # brief 检查清单文件
    # private
    def check_manifest(self):
        # 获取清单文件中的excel文件名
        self._excel_set = set()
        for line in self._manifest_file:
            self._excel_set.add(line.strip())

    # brief 通过目录生成文件
    # private
    def generate_def_by_folder(self):
        for dir_path, dir_names, file_names in os.walk(XLSX_FOLDER):
            # 过滤掉清单中不存在的文件
            valid_filename_list = list(filter(self.file_filter, file_names))
            # 生成数据文件
            folder = dir_path.replace('\\', '/')
            for filename in valid_filename_list:
                self.generate_def_by_path(f"{folder}/{filename}")

    # brief 通过路径生成数据文件
    # private
    def generate_def_by_path(self, filepath):
        # 工作簿实例
        workbook = load_workbook(filepath, data_only=True)
        if not workbook:
            print(f"工作簿加载异常:{filepath}")
            return
        sheet_name_list = workbook.get_sheet_names()
        if not sheet_name_list:
            print(f"找不到任何表格:{filepath}")
            return
        # 清单表
        manifest_sheet: Worksheet = workbook.get_sheet_by_name(sheet_name_list[0])
        # 遍历清单工作簿
        for i in range(manifest_sheet.max_row):
            # 清单第一列记录需要生成数据的表格
            sheet_name = str(manifest_sheet.cell(column=1, row=i + 1).value)
            fix_sheet_name = str_util.fix_str(sheet_name)
            if not fix_sheet_name or fix_sheet_name == "":
                print(f"找不到sheet名:{filepath} {fix_sheet_name}")
                continue
            # 数据表
            data_sheet = workbook.get_sheet_by_name(sheet_name)
            if not data_sheet:
                print(f"sheet不存在:{filepath} {fix_sheet_name}")
                continue
            # 数据文件路径
            output_path = f"{DEF_TARGET_FOLDER}/Def{sheet_name}.cs"
            self.generate_def_by_sheet(data_sheet, output_path, filepath)

    # brief 通过表名生成数据文件
    # private
    def generate_def_by_sheet(self, data_sheet: Worksheet, output_path, filepath):
        def_col = self.find_def_col(data_sheet)
        # 定义不存在
        if def_col == -1:
            return
        print(output_path)
        def_file = codecs.open(output_path, "w", "utf-8")
        # 写入标题
        class_name = f"Def{data_sheet.title}"
        self.write_header_str(def_file, class_name, filepath)
        # 写入代码
        self.write_code(def_file, class_name, data_sheet, def_col)
        def_file.close()

    # brief 写入标题字符
    # private
    @staticmethod
    def write_header_str(script_file, class_name, filepath):
        if not isinstance(class_name, str):
            print(f"文件名格式错误: {class_name}")
            return
        if not script_file:
            return
        str_list = [
            "// ******************************************************************",
            f"//       /\\ /|       @file       {class_name}.cs",
            f"//       \\ V/        @brief      excel(由python自动生成) {filepath}",
            "//       | \"\")       @author     Shadowrabbit, yue.wang04@mihoyo.com",
            "//       /  |",
            "//      /  \\\\        @Modified   2022-04-25 13:25:11",
            "//    *(__\\_\\        @Copyright  Copyright (c)  2022, Shadowrabbit",
            "// ******************************************************************"
        ]
        result = os.linesep.join(str_list)
        script_file.write(result)

    # brief 写入
    # private
    def write_code(self, def_file, class_name: str, data_sheet: Worksheet, def_col: int):
        def_file.write(os.linesep)
        def_file.write(os.linesep)
        def_file.write(f"namespace {config_def.NAME_SPACE_NAME}")
        def_file.write(os.linesep)
        def_file.write("{")
        # 写入
        self.write_def(def_file, class_name, data_sheet, def_col)
        def_file.write(os.linesep)
        def_file.write("}")

    # brief 写入
    # private
    def write_def(self, def_file, class_name: str, data_sheet: Worksheet, def_col: int):
        if not def_file:
            return
        def_file.write(os.linesep)
        def_file.write(f"    public static class {class_name}")
        def_file.write(os.linesep)
        def_file.write("    {")
        for col in range(data_sheet.max_column):
            # 不是所在列
            if col + 1 != def_col:
                continue
            # 查找注释所在列
            annotate_col = self.find_annotate_col(data_sheet)
            # 当前的id
            i = 1
            for row in range(data_sheet.max_row):
                # 前三行忽略
                if row < 3:
                    continue
                cfg_id = data_sheet.cell(column=1, row=row + 1).value
                # 没填id
                if not cfg_id:
                    continue
                def_file.write(os.linesep)
                suffix = ""
                # 存在注释的情况 并且有id
                if annotate_col != -1:
                    annotate_value = data_sheet.cell(column=annotate_col, row=row + 1).value
                    if annotate_value:
                        suffix = f" //{str(data_sheet.cell(column=annotate_col, row=row + 1).value)}"
                # 定义
                def_value = data_sheet.cell(column=def_col, row=row + 1).value
                # 没填命名 用脚本+行数默认命名
                if not def_value:
                    def_value = f"D{cfg_id}"
                def_name = to_big_camel_case(str(def_value))
                def_file.write(f"        public static readonly string {def_name} = \"{cfg_id}\"; {suffix}")
                i += 1
        def_file.write(os.linesep)
        def_file.write("    }")

    # brief 查找定义所在列
    # private
    @staticmethod
    def find_def_col(data_sheet: Worksheet):
        for col in range(data_sheet.max_column):
            # 没有列定义
            if str(data_sheet.cell(column=col + 1, row=1).value) == DEF_MARK:
                return col + 1
        return -1

    # brief 查找注释所在列
    # private
    @staticmethod
    def find_annotate_col(data_sheet: Worksheet):
        for col in range(data_sheet.max_column):
            # 不是注释列
            if str(data_sheet.cell(column=col + 1, row=1).value) == ANNOTATE_MARK:
                return col + 1
        return -1

    # brief 文件名是否在清单中
    # private
    def file_filter(self, filename) -> bool:
        return filename in self._excel_set
